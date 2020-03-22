from pprint import pprint
from typing import OrderedDict
import openpyxl
from openpyxl.utils.cell import coordinate_from_string

from pathlib import Path
import logging

logging.basicConfig(level=logging.DEBUG)
# CURRENT ERROR INDEX 13

class XlsxParaser():
    def __init__(self, aPath="", aLog_level = "DEBUG"):
        self.path = aPath    # xlsx path
        self.workbook = None # openpyxl workbook
        self.sheet = None    # current sheet
        self.columns = []    # class Column

        logging.getLogger().setLevel(self.__set_log_level(aLog_level))

    #-------------------------------------------------------------------------------#

    def open(self, aDefault_name = "Sheet1") -> bool:
        """
            aDefault_name: str -> default sheet name
        """

        if not self.path:
            logging.error("Path is empty, error num: 1")
            return False

        oWorkbook_path = Path(self.path)
        logging.debug(f"Xlsx path: {oWorkbook_path}.")

        if oWorkbook_path.is_file(): # open workbook
            self.workbook = openpyxl.load_workbook(str(oWorkbook_path))
            self.sheet = self.workbook.active
            aDefault_name = self.sheet.title
            logging.debug("Opened Workbook.")
        else:                        # create new workbook
            self.workbook = openpyxl.Workbook()
            self.sheet = self.workbook.active
            self.sheet.title = aDefault_name
            logging.debug("Created Workbook.")
        
        logging.debug(f"Current sheet: {aDefault_name}.")
        return True

    #-------------------------------------------------------------------------------#

    def close(self) -> None:
        self.workbook.save(str(self.path))
        logging.debug(f"Saving workbook to {self.path}.")

    #-------------------------------------------------------------------------------#
    # SHEETS
    #-------------------------------------------------------------------------------#

    def get_sheets(self) -> list:
        return self.workbook.sheetnames

    #-------------------------------------------------------------------------------#

    def set_sheet(self, aSheet_name) -> None:
        """
            aSheet_name: str -> name of the sheet to be set
        """

        xSheets = self.get_sheets()
        if aSheet_name not in xSheets:
            logging.error(f"Error sheet: {aSheet_name} not found, error num: 2")
            raise RuntimeError("No sheet found")

        for nSheet in range(len(self.workbook.sheetnames)):
            if self.workbook.sheetnames[nSheet] == aSheet_name:
                self.workbook.active = nSheet
                self.sheet = self.workbook.active
                logging.debug(f"Current sheet: {aSheet_name}")
                return

    #-------------------------------------------------------------------------------#

    def create_sheet(self, aSheet_name) -> None:
        """
            aSheet_name: str -> new sheet name
        """

        self.workbook.create_sheet(aSheet_name)
        self.set_sheet(aSheet_name)
        logging.debug(f"Created sheet: {aSheet_name}")

    #-------------------------------------------------------------------------------#

    def delete_sheet(self, aSheet_name) -> None:
        """
            aSheet_name: str -> sheet to be deleted
        """

        try:
            aSheet = self.workbook[aSheet_name]
        except Exception as e:
            raise RuntimeError("No sheet found")
        self.workbook.remove(aSheet)
        logging.debug(f"Removed sheet: {aSheet_name}")

    #-------------------------------------------------------------------------------#

    def rename_sheet(self, aSheet_old, aSheet_new) -> None:
        """
            aSheet_old: str -> old sheet name
            aSheet_new: str -> new sheet name
        """

        xSheets = self.get_sheets()
        if aSheet_old not in xSheets:
            logging.error(f"Error renaming, sheet: {aSheet_old} not found, error num: 3")
            raise RuntimeError("No sheet found")

        for nSheet in range(len(self.workbook.sheetnames)):
            if self.workbook.sheetnames[nSheet] == aSheet_old:
                self.workbook[aSheet_old].title = aSheet_new
                self.set_sheet(aSheet_new)
                return

    #-------------------------------------------------------------------------------#
    # HEADERS
    #-------------------------------------------------------------------------------#

    def get_headers(self) -> list:
        return [oCol.header for oCol in self.columns]

    #-------------------------------------------------------------------------------#

    def find_headers(self, aSheet_name = "") -> bool:
        """
            aSheet_name: str  -> sheet name
        """

        xColumns = []

        if aSheet_name:
            self.set_sheet(aSheet_name)
        else:
            aSheet_name = self.sheet.title

        # Find the first cell with values in each column.
        for oCol in self.sheet.iter_cols(max_row=50, max_col=50):
            for oCell in oCol:
                if not oCell.value:
                    continue

                aHeader = oCell.value
                aIndex = coordinate_from_string(oCell.coordinate)[0]
                nStart = coordinate_from_string(oCell.coordinate)[1]

                oColumn = Column(aSheet_name, aHeader, aIndex, nStart, 0)
                xColumns.append(oColumn)

                break

        # If no columns found return false
        if not xColumns:
            return False

        # Add new columns
        self.columns.extend(xColumns)

        # Find the last cell with value in each column.
        for oColumn in self.columns:
            nSheet_end = self.sheet.max_row
            while True:
                oCell = f'{oColumn.index}{nSheet_end}'
                if self.sheet[oCell].value or nSheet_end == oColumn.start:
                    oColumn.end = nSheet_end
                    logging.debug(f"Found Column: {oColumn}")
                    break
                nSheet_end -= 1
        
        return True

    #-------------------------------------------------------------------------------#

    def set_headers(self, xData, aSheet_name = ""):
        """
            xData: list
                oData: dict
                    header: str  -> name of the column
                    index : str  -> column letter
                    start : str  -> column start
            aSheet_name: str  -> sheet name
        """

        if not self.__check_args({"header", "index", "start"}, xData):
            logging.error(f"Parameters are missing some keys, error num: 4")
            return False

        if aSheet_name:
            self.set_sheet(aSheet_name)

        for oHeader in xData:
            xTo_insert = []

            aHeader = oHeader["header"]
            aIndex = oHeader["index"]
            nStart = oHeader["start"]
            nEnd = nStart

            oColumn = Column(aSheet_name, aHeader, aIndex, nStart, nEnd)
            self.columns.append(oColumn)
            
            xTo_insert.append({"header": aHeader, "data": aHeader, "row": oColumn.start})

            self.append_rows(xTo_insert, aSheet_name=aSheet_name, bAppend_if_none=False)

            logging.debug(f"Created Column: {oColumn}")

    #-------------------------------------------------------------------------------#
    # ROWS
    #-------------------------------------------------------------------------------#
                 
    def get_rows(self, oData, aSheet_name = None) -> list:
        """
        oData: dict
            row   : int  -> the row to be returned
            OR
            header: str  -> header name
            search: str  -> cell value
        aSheet_name: str -> the name of the sheet for the data insertion
        """

        if aSheet_name:
            self.set_sheet(aSheet_name)

        if not oData.get("row") and (not oData.get("header") or not oData.get("search")):
            logging.error("Invalid parameters, error num: 5")
            raise RuntimeError()

        if oData.get("row"):
            return self.__get_row_by_number(oData.get("row"))

        elif oData.get("header") and oData.get("search"):
            return self.__get_rows_by_string(oData)

        else:
            logging.error("Not implemented, error num: 12")
            raise RuntimeError()
    
    #-------------------------------------------------------------------------------#
          
    def append_rows(self, xData, aSheet_name = None, bAppend_if_none = True) -> bool:
        """
        xData: list
            oData: dict
                header: str       -> header name
                data  : str       -> cell value
                row   : str (opt) -> row, if present insert at that
                                     position else insert at the end
        aSheet_name: str          -> the name of the sheet for the data insertion
        bAppend_if_none: bool     -> if true move the columns last pointer
                                     by 1 even if the data in that column is empty
        """

        if aSheet_name:
            self.set_sheet(aSheet_name)

        if not self.__check_args({"header", "data"}, xData):
            logging.error(f"Parameters are missing some keys, error num: 6")
            return False

        for oColumn in self.columns:
            if aSheet_name and oColumn.sheet_name != aSheet_name:
                continue
            bData_added = False
            for oData in xData:
                if oData["header"] == oColumn.header:
                    if oData.get("row"):
                        self.sheet[oColumn.get_by_pos(oData["row"])].value = oData["data"]
                    else:
                        self.sheet[oColumn.get_last()].value = oData["data"]
                    bData_added = True
                    break
                
            if not bData_added and bAppend_if_none:
                oColumn.get_last()

        return True

    #-------------------------------------------------------------------------------#

    def update_rows(self, xData, aSheet_name = None, insert_if_not_found = True) -> bool:
        """
        xData: list
            oData: dict:
                header  : str      -> header name
                old_data: str      -> old value
                new_data: str      -> new value
        aSheet_name: str           -> the name of the sheet for the data insertion
        insert_if_not_found: bool  -> if true insert the row if the old_data wasnt found
        """

        if aSheet_name:
            self.set_sheet(aSheet_name)

        if not self.__check_args({"header", "old_data", "new_data"}, xData):
            logging.error(f"Parameters are missing some keys, error num: 9")
            return False

        for oData in xData:
            oColumn = self.__get_column(oData["header"])
            if not oColumn:
                logging.error(f'Column {oData["header"]} not found, error num: 10')
                return False

            bUpdated = False
            nCurrent = oColumn.start
            while nCurrent <= oColumn.end:
                oCell = self.sheet[oColumn.get_by_pos(nCurrent)]
                if oCell.value == oData["old_data"]:
                    oCell.value = oData["new_data"]
                    bUpdated = True
            
                nCurrent += 1

            if not bUpdated and insert_if_not_found:
                if not self.append_rows([{"header": oData["header"], "data": oData["new_data"]}]):
                    return False
        
        return True

    #-------------------------------------------------------------------------------#

    def remove_rows(self, xRows, aSheet_name = None) -> bool:
        """
            xRows: list -> list of rows numbers
            aSheet_name: str           -> the name of the sheet for the data insertion
        """

        if aSheet_name:
            self.set_sheet(aSheet_name)

        if not isinstance(xRows, list):
            logging.error("Parameter xRows must be type list, error num: 11")
            return False

        for nRow in sorted(xRows, reverse=True):
            self.sheet.delete_rows(nRow)

        return True

    #-------------------------------------------------------------------------------#
    # PRIVATE METHODS
    #-------------------------------------------------------------------------------#

    def __get_column(self, aHeader):
        for oCol in self.columns:
            if oCol.header == aHeader:
                return oCol
        return None

    #-------------------------------------------------------------------------------#

    def __set_log_level(self, aLog_level):
        if aLog_level == "DEBUG":
            return logging.DEBUG
        elif aLog_level == "INFO":
            return logging.INFO
        elif aLog_level == "WARNING":
            return logging.WARNING
        elif aLog_level == "ERROR":
            return logging.ERROR
        elif aLog_level == "CRITICAL":
            return logging.CRITICAL
        elif aLog_level == "NOTSET":
            return logging.NOTSET
        else:
            return logging.DEBUG

    #-------------------------------------------------------------------------------#

    def __check_args(self, oRequired, xData) -> bool:
        """
            Check that all the dicts in xData contain the strings in oRequired set
            oRequired: set  -> set of required keys
            xData    : list -> list of data dicts
        """
        if not all(len(set(oData.keys()).intersection(oRequired)) >= len(oRequired) for oData in xData):
            return False
        return True

    #-------------------------------------------------------------------------------#

    def __get_row_by_number(self, nRow) -> list:
        """
            nRow: int -> row number
        """

        oResult = self.__get_result_dict()

        for oColumn in self.columns:
            if nRow < 0 or nRow < oColumn.start or nRow > oColumn.end:
                continue

            oCell = self.sheet[oColumn.get_by_pos(nRow)]
            oResult["data"].update({oColumn.header: oCell.value})
            if not oResult["row"]:
                oResult["row"] = nRow

        return [oResult]

    #-------------------------------------------------------------------------------#

    def __get_rows_by_string(self, oData) -> list:
        """
            oData: dict
                header: str  -> header name
                search: str  -> cell value
        """

        xResult = []

        oColumn = self.__get_column(oData["header"])
        if not oColumn:
            return [self.__get_result_dict()]

        nCurrent = oColumn.start
        while nCurrent <= oColumn.end:
            oCell = self.sheet[oColumn.get_by_pos(nCurrent)]
            if oCell.value == oData["search"]:
                xResult.extend(self.__get_row_by_number(nCurrent))
            nCurrent += 1

        if xResult:
            return xResult
        else:
            return [self.__get_result_dict()]

    #-------------------------------------------------------------------------------#

    def __get_result_dict(self):
        return {
            "row": None,
            "data": {}
        }


class Column():
    def __init__(self, aSheet="", aHeader="", aIndex="", nStart=0, nEnd=0):
        self.sheet_name = aSheet # column sheet name
        self.header = aHeader    # column header name
        self.index = aIndex      # column index, eg: A,B,C,...
        self.start = nStart      # column start
        self.end = nEnd          # column end

    #-------------------------------------------------------------------------------#

    def get_last(self):
        self.end += 1
        return f'{self.index}{self.end}'

    #-------------------------------------------------------------------------------#

    def get_by_pos(self, nPos):
        """
            nPos: int -> the position in the column
        """
        nPos = int(nPos)
        if nPos > self.end:
            self.end = nPos
        return f'{self.index}{nPos}'

    #-------------------------------------------------------------------------------#

    def __repr__(self):
        return f'Sheet: {self.sheet_name}, Header: {self.header}, index: {self.index}, start: {self.start}, end: {self.end}'
    
    #-------------------------------------------------------------------------------#
    
    def __str__(self):
        return f'Sheet: {self.sheet_name}, Header: {self.header}, index: {self.index}, start: {self.start}, end: {self.end}'
